from instagrapi import Client
from instagrapi.exceptions import LoginRequired
from datetime import datetime, timedelta
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook


class WorkbookManipulation:
    def __init__(self, filename, sheet_name_1, sheet_name_2, sheet_name_3):
        self.filename = filename
        self.sheet_name_1 = sheet_name_1
        self.sheet_name_2 = sheet_name_2
        self.sheet_name_3 = sheet_name_3
        self.wb = None

    def check_file_exists(self):
        if os.path.exists(self.filename):
            self.wb = load_workbook(self.filename)
            print(f"File '{self.filename}' already exists.")
        else:
            self.wb = Workbook()
            self.wb.create_sheet(title=self.sheet_name_1)
            self.wb.create_sheet(title=self.sheet_name_2)
            self.wb.create_sheet(title=self.sheet_name_3)

            self.setup_headers(self.sheet_name_2)
            self.setup_headers(self.sheet_name_3)

            if "Sheet" in self.wb.sheetnames:
                self.wb.remove(self.wb["Sheet"])

            self.wb.save(file_name)
            print(f"File '{self.filename}' was created")

    def get_users_from_sheet1(self):
        df = pd.read_excel(self.filename, sheet_name=self.sheet_name_1, engine='openpyxl')
        if df.empty:
            print("User Sheet is empty, please update it with data")
            exit(1)
        target_usernames_list = df.iloc[:, 0].tolist()
        print("[*] Target list length:", len(target_usernames_list))
        if not target_usernames_list:
            exit(1)
        return target_usernames_list

    def sheet_fill_data(self, sheet_name, df):
        sheet = self.wb[sheet_name]

        df = df.dropna(how='all')

        data = df.values.tolist()
        for row in data:
            sheet.append(row)

        sheet.auto_filter.ref = sheet.dimensions

    def save_wb(self):
        self.wb.save(self.filename)

    def setup_headers(self, sheet_name):
        sheet = self.wb[sheet_name]
        headers = ["user_id", "target_username", "link", "datetime"]

        if sheet.max_row == 1:
            for col_index, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_index, value=header)

        sheet.freeze_panes = sheet['A2']


class ScraperInstagram:
    def __init__(self, username, password, session_file="./instagram_session.json"):
        self.username = username
        self.password = password
        self.session_file = session_file
        self.cl = Client()
        self.one_week_ago = datetime.now() - timedelta(days=7)
        self.two_hours_ago = datetime.now() - timedelta(hours=10)

    def login_with_session(self):
        try:
            self.cl.load_settings(self.session_file)
            self.cl.login(self.username, self.password)
        except FileNotFoundError:
            print("Session was not found. Auth...")
            self.cl.login(self.username, self.password)
            self.cl.dump_settings(self.session_file)
        except LoginRequired:
            if os.path.exists(session_file):
                os.remove(session_file)
            print("Session is not valid. Auth again...")
            self.cl.login(self.username, self.password)
            self.cl.dump_settings(self.session_file)

    def scraping_process(self, target_username_list):
        posts_data = []
        stories_data = []
        for target_username in target_username_list:
            try:
                user_id = self.cl.user_id_from_username(target_username)
                print("-" * 60)
                print(f"User id @{target_username}: {user_id}")

                # take all media (posts + reels) and stories
                media = self.cl.user_medias(user_id, amount=50)
                stories = self.cl.user_stories(user_id)

                # filtration by the date
                recent_posts = self.filter_by_date(media, self.one_week_ago)
                recent_stories = self.filter_by_date(stories, self.two_hours_ago)

                # extend data with the parsed information
                posts_data.extend(self.parse_media_data(recent_posts, "media", user_id, target_username))
                stories_data.extend(self.parse_media_data(recent_stories, "story", user_id, target_username))

            except Exception as e:
                print(f"Error during parsing the @{target_username}: {e}")
        return posts_data, stories_data

    @staticmethod
    def parse_media_data(data, type_media, user_id, target_username):
        data_list = []
        if data:
            for i, el in enumerate(data, start=1):
                url = None
                if type_media == "media":
                    url = f"https://www.instagram.com/p/{el.code}/" if el.code else "Ссылка недоступна"
                elif type_media == "story":
                    url = f"https://www.instagram.com/stories/{target_username}/{el.pk}/"

                formatted_date = el.taken_at.strftime('%Y-%m-%d-%H-%M')

                print(f"\n******************** {type_media} #{i} ********************")
                print(f"ID of {type_media}: {el.id}")
                print(f"Date of publication: {formatted_date}")
                print(f"URL of {type_media}: {url}")
                # print(f"Type: {post.media_type}")
                # print(f"Description: {post.caption_text}")
                # print(f"URL of the content: {post.thumbnail_url if post.media_type == 1 else post.video_url}")
                # print(f"Number of likes: {post.like_count}")
                # print(f"Number of comments: {post.comment_count}")

                current_data = {
                    "user_id": user_id,
                    "username": target_username,
                    "link": url,
                    "datetime": formatted_date,
                }
                data_list.append(current_data)
        else:
            print(f"User @{target_username} does not have {type_media} data.")

        return data_list

    @staticmethod
    def filter_by_date(data, timedelta):
        filtered_data = [
            el for el in data
            if datetime.fromtimestamp(el.taken_at.timestamp()) > timedelta
        ]
        return filtered_data


if __name__ == "__main__":
    file_name = "instagram_scraper.xlsx"
    sheet_name_1 = "UserList"
    sheet_name_2 = "Posts"
    sheet_name_3 = "Stories"
    workbook = WorkbookManipulation(file_name, sheet_name_1, sheet_name_2, sheet_name_3)
    workbook.check_file_exists()
    target_username_list = workbook.get_users_from_sheet1()

    username = "USERNAME"
    password = "PASSWORD"
    session_file = "./instagram_session.json"
    scraper = ScraperInstagram(username, password, session_file)
    scraper.login_with_session()

    posts_data, stories_data = scraper.scraping_process(target_username_list)
    posts_df = pd.DataFrame(posts_data)
    stories_df = pd.DataFrame(stories_data)
    workbook.sheet_fill_data(sheet_name_2, posts_df)
    workbook.sheet_fill_data(sheet_name_3, stories_df)
    workbook.save_wb()
    print(f"[*] File '{file_name}' was saved successfully!")
